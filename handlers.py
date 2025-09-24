# ================== handlers.py ==================
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram import KeyboardButton, ReplyKeyboardMarkup
from telegram.constants import ChatAction
from telegram.ext import ContextTypes, ConversationHandler
from sheet import sheet_handler
from difflib import get_close_matches
import re
import time
import os
from datetime import datetime
from collections import Counter
import asyncio
import ast
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from collections import Counter
from openpyxl.chart import LineChart

WHITELIST_FILE = "whitelist.py"

def load_whitelist():
    try:
        with open(WHITELIST_FILE, "r") as f:
            data = f.read().strip()
            if data.startswith("WHITELIST"):
                wl = ast.literal_eval(data.split("=", 1)[1].strip())
                return wl
    except FileNotFoundError:
        return []
    return []

# Load known brands from external file
with open("brands.txt", "r", encoding="utf-8") as f:
    KNOWN_BRANDS = [line.strip() for line in f if line.strip()]

STORE_LOCATION_URL = "https://maps.app.goo.gl/RiURGuNtMyzSCmyDA"

CHOOSE_CATEGORY, ASK_MODEL = range(2)

inventory_cache = {
    "timestamp": 0,
    "data": None
}

CACHE_DURATION = 300

LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)
STATS_DIR = "stats"
os.makedirs(STATS_DIR, exist_ok=True)
CATEGORIES = ["LCD", "Battery", "Connector", "Glass", "COVER", "SERSOU"]

CATEGORY_MAPPING = {
    "LCD": "LCD",
    "Battery": "BATTERIE",
    "Connector": "CC",
    "Glass": "GLASS",
    "COVER": "COVER",
    "SERSOU": "SERSOU"
}

async def log_request(category, model, available):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    excel_file = os.path.join(STATS_DIR, f"{date_str}.xlsx")

    # Create file if not exists
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Search Log"
        ws1.append(["Time", "Category", "Model", "Status"])

        wb.create_sheet("Category Summary")
        wb.create_sheet("Not Available")
        wb.create_sheet("Charts")
        wb.create_sheet("Daily Trends")

        wb.save(excel_file)

    # Load workbook
    wb = load_workbook(excel_file)

    ws_log = wb["Search Log"]
    ws_summary = wb["Category Summary"]
    ws_na = wb["Not Available"]
    ws_charts = wb["Charts"]
    ws_trends = wb["Daily Trends"]

    # Append to search log
    status = "Available" if available else "Not available"
    ws_log.append([now.strftime("%H:%M:%S"), category, model, status])

    # Rebuild Not Available
    ws_na.delete_rows(1, ws_na.max_row)
    ws_na.append(["Time", "Category", "Model", "Status"])
    for row in ws_log.iter_rows(min_row=2, values_only=True):
        if row[3] == "Not available":
            ws_na.append(row)

    # Rebuild Category Summary
    ws_summary.delete_rows(1, ws_summary.max_row)
    ws_summary.append(["Category", "Searches"])
    counts = {cat: 0 for cat in CATEGORIES}
    for row in ws_log.iter_rows(min_row=2, values_only=True):
        if row[1] in counts:
            counts[row[1]] += 1
    for cat in CATEGORIES:
        ws_summary.append([cat, counts[cat]])

    # Rebuild Daily Trends
    ws_trends.delete_rows(1, ws_trends.max_row)
    ws_trends.append(["Hour", "Searches"])
    hours = [row[0][:2] for row in ws_log.iter_rows(min_row=2, values_only=True)]
    hour_counts = Counter(hours)
    for h in range(24):
        hour_str = f"{h:02d}"
        ws_trends.append([hour_str, hour_counts.get(hour_str, 0)])

    # Rebuild Charts
    ws_charts._charts.clear()

    # Pie chart
    pie = PieChart()
    data = Reference(ws_summary, min_col=2, min_row=1, max_row=1+len(CATEGORIES))
    labels = Reference(ws_summary, min_col=1, min_row=2, max_row=1+len(CATEGORIES))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Searches by Category"
    ws_charts.add_chart(pie, "B2")

    # Bar chart
    available_count = sum(1 for row in ws_log.iter_rows(min_row=2, values_only=True) if row[3] == "Available")
    not_available_count = sum(1 for row in ws_log.iter_rows(min_row=2, values_only=True) if row[3] == "Not available")

    ws_charts["E1"] = "Status"
    ws_charts["F1"] = "Count"
    ws_charts["E2"] = "Available"
    ws_charts["F2"] = available_count
    ws_charts["E3"] = "Not available"
    ws_charts["F3"] = not_available_count

    bar = BarChart()
    data = Reference(ws_charts, min_col=6, min_row=1, max_row=3)
    labels = Reference(ws_charts, min_col=5, min_row=2, max_row=3)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = "Availability Status"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Status"
    ws_charts.add_chart(bar, "E5")

    # Line chart (Daily Trends)
    line = LineChart()
    data = Reference(ws_trends, min_col=2, min_row=1, max_row=25)
    labels = Reference(ws_trends, min_col=1, min_row=2, max_row=25)
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.title = "Searches by Hour"
    line.x_axis.title = "Hour"
    line.y_axis.title = "Searches"
    ws_charts.add_chart(line, "B20")

    wb.save(excel_file)

async def get_cached_inventory():
    now = time.time()
    if inventory_cache["data"] is None or now - inventory_cache["timestamp"] > CACHE_DURATION:
        data = await sheet_handler.get_inventory()
        for row in data:
            try:
                qt_raw = str(row.get("QT", "")).strip()
                row["QT"] = float(qt_raw) if qt_raw and float(qt_raw) > 0 else 0
                pu_raw = str(row.get("PU", "")).strip()
                row["PU"] = float(pu_raw) if pu_raw else 0
            except:
                row["QT"] = 0
                row["PU"] = 0
        inventory_cache["data"] = data
        inventory_cache["timestamp"] = now
    return inventory_cache["data"]

import asyncio


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    
    user_id = update.effective_user.id
    whitelist = load_whitelist()

    if user_id not in whitelist:
        if update.message:
            await update.message.reply_text("❌ Désolé vous étes pas autorisé a utilisé ce bot.")
        elif update.callback_query:
            await update.callback_query.message.reply_text("❌ Désolé vous étes pas autorisé a utilisé ce bot.")
        return ConversationHandler.END  # stop the conversation flow

    
    context.user_data.clear()
    buttons = [
        [InlineKeyboardButton("📱 LCD", callback_data="LCD")],
        [InlineKeyboardButton("🔋 Batterie", callback_data="Battery")],
        [InlineKeyboardButton("🔌 Connecteur", callback_data="Connector")],
        [InlineKeyboardButton("🧿 Glass", callback_data="Glass")],
        [InlineKeyboardButton("📦 Cover", callback_data="COVER")],
        [InlineKeyboardButton("🛠 Sersou", callback_data="SERSOU")]
    ]
    markup = InlineKeyboardMarkup(buttons)

    # Send "processing..." first
    if update.message:
        processing_msg = await update.message.reply_text("⏳ Processing...")
    elif update.callback_query:
        processing_msg = await update.callback_query.message.reply_text("⏳ Processing...")
    else:
        return CHOOSE_CATEGORY

    async def send_menu():
        if update.message:
            await processing_msg.edit_text("Bienvenue ! Choisissez une catégorie :", reply_markup=markup)
        elif update.callback_query:
            await processing_msg.edit_text("Bienvenue ! Choisissez une catégorie :", reply_markup=markup)

    try:
        await asyncio.wait_for(send_menu(), timeout=5)
    except asyncio.TimeoutError:
        await processing_msg.edit_text(
            "❌ Une erreur est survenue lors du traitement de votre requête. Veuillez réessayer.",
            reply_markup=markup
        )

    return CHOOSE_CATEGORY


async def category_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    try:
        await query.answer()
    except Exception:
        pass
    category = query.data

    if category == "restart":
        return await restart_search(update, context)

    context.user_data["category"] = category

    # Show instant "processing…" message
    processing_msg = await query.message.reply_text("⏳ Processing...")

    # Timeout fallback (5s)
    async def timeout_fallback():
        await asyncio.sleep(5)
        if not context.user_data.get("response_sent"):
            buttons = [
                [InlineKeyboardButton("📱 LCD", callback_data="LCD")],
                [InlineKeyboardButton("🔋 Batterie", callback_data="Battery")],
                [InlineKeyboardButton("🔌 Connecteur", callback_data="Connector")],
                [InlineKeyboardButton("🧿 Glass", callback_data="Glass")],
                [InlineKeyboardButton("📦 Cover", callback_data="COVER")],
                [InlineKeyboardButton("🛠 Sersou", callback_data="SERSOU")]
            ]
            markup = InlineKeyboardMarkup(buttons)
            await query.message.reply_text(
                "❌ An error has occurred while processing your request. Please try again.",
                reply_markup=markup
            )

    timeout_task = asyncio.create_task(timeout_fallback())

    try:
        await asyncio.sleep(0.5)

        timeout_task.cancel()
        await processing_msg.delete()
        context.user_data["response_sent"] = True

        await query.message.reply_text(
            f"Vous avez choisi la catégorie : {category}\n\nVeuillez maintenant entrer le modèle de téléphone 📱:"
        )
        return ASK_MODEL

    except Exception as e:
        timeout_task.cancel()
        await processing_msg.edit_text(
            f"❌ Une erreur est survenue ({e}). Veuillez réessayer."
        )
        return CHOOSE_CATEGORY

async def handle_model(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_input = update.message.text.strip().lower()
    category = context.user_data.get("category")
    if not category:
        await update.message.reply_text("❌ Veuillez d'abord choisir une catégorie avec /start.")
        return ConversationHandler.END

    # Show instant feedback
    processing_msg = await update.message.reply_text("⏳ Processing...")

    # Start timeout task
    async def timeout_fallback():
        await asyncio.sleep(5)
        if not context.user_data.get("response_sent"):
            buttons = [
                [InlineKeyboardButton("📱 LCD", callback_data="LCD")],
                [InlineKeyboardButton("🔋 Batterie", callback_data="Battery")],
                [InlineKeyboardButton("🔌 Connecteur", callback_data="Connector")],
                [InlineKeyboardButton("🧿 Glass", callback_data="Glass")],
                [InlineKeyboardButton("📦 Cover", callback_data="COVER")],
                [InlineKeyboardButton("🛠 Sersou", callback_data="SERSOU")]
            ]
            markup = InlineKeyboardMarkup(buttons)
            await update.message.reply_text(
                "❌ An error has occurred while processing your request. Please try again.",
                reply_markup=markup
            )

    timeout_task = asyncio.create_task(timeout_fallback())

    await update.message.chat.send_action(action=ChatAction.TYPING)

    if not re.match(r"^[a-z0-9\s\-+_.]+$", user_input):
        timeout_task.cancel()
        await processing_msg.delete()
        await update.message.reply_text("❌ Entrée invalide. Veuillez saisir un modèle de téléphone valide.")
        return ASK_MODEL

    # Match brand correction
    input_words = user_input.split()
    possible_brand = input_words[0]
    matched_brand = get_close_matches(possible_brand, [b.lower() for b in KNOWN_BRANDS], n=1, cutoff=0.8)
    if matched_brand:
        corrected_brand = matched_brand[0]
        user_input = user_input.replace(possible_brand, corrected_brand)

    designation_keyword = CATEGORY_MAPPING.get(category, category)
    data = await get_cached_inventory()
    filtered_data = [row for row in data if designation_keyword.lower() in row["Designation1"].lower()]

    potential_matches = [row for row in filtered_data if user_input in row["Designation1"].lower()]

    if not potential_matches:
        all_model_names = [row["Designation1"].lower() for row in filtered_data]
        close = get_close_matches(user_input, all_model_names, n=1, cutoff=0.9)
        timeout_task.cancel()
        await processing_msg.delete()
        if close:
            for row in filtered_data:
                if close[0] == row["Designation1"].lower():
                    context.user_data["response_sent"] = True
                    return await respond_with_inventory_info(update, context, row, category, user_input)

        await log_request(category, user_input, False)
        await update.message.reply_text(
            f"❌ {user_input} n'est pas disponible dans notre inventaire. Essayez un autre modèle ou recommencez."
        )
        return await start(update, context)

    timeout_task.cancel()
    await processing_msg.delete()
    context.user_data["response_sent"] = True

    if len(potential_matches) == 1:
        return await respond_with_inventory_info(update, context, potential_matches[0], category, user_input)

    # Multiple matches found
    context.user_data['pending_matches'] = potential_matches
    context.user_data['search_query'] = user_input
    buttons = [[InlineKeyboardButton(row['Designation1'], callback_data=f"select::{row['Designation1']}")] for row in potential_matches]
    markup = InlineKeyboardMarkup(buttons)
    await update.message.reply_text("🔍 Plusieurs correspondances trouvées. Veuillez préciser :", reply_markup=markup)
    return ASK_MODEL

async def handle_model_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Show "processing…" immediately
    processing_msg = await query.message.reply_text("⏳ Processing...")

    async def timeout_fallback():
        await asyncio.sleep(5)
        if not context.user_data.get("response_sent"):
            buttons = [
                [InlineKeyboardButton("📱 LCD", callback_data="LCD")],
                [InlineKeyboardButton("🔋 Batterie", callback_data="Battery")],
                [InlineKeyboardButton("🔌 Connecteur", callback_data="Connector")],
                [InlineKeyboardButton("🧿 Glass", callback_data="Glass")],
                [InlineKeyboardButton("📦 Cover", callback_data="COVER")],
                [InlineKeyboardButton("🛠 Sersou", callback_data="SERSOU")]
            ]
            markup = InlineKeyboardMarkup(buttons)
            await query.message.reply_text(
                "❌ An error has occurred while processing your request. Please try again.",
                reply_markup=markup
            )

    timeout_task = asyncio.create_task(timeout_fallback())

    try:
        selected_model = query.data.split("::")[1]
        category = context.user_data.get("category")

        if not category:
            timeout_task.cancel()
            await processing_msg.edit_text("❌ La catégorie est introuvable. Veuillez redémarrer avec /start.")
            return ConversationHandler.END

        data = await get_cached_inventory()
        for row in data:
            if row['Designation1'] == selected_model:
                timeout_task.cancel()
                await processing_msg.delete()
                context.user_data["response_sent"] = True
                return await respond_with_inventory_info(query, context, row, category, selected_model)

        timeout_task.cancel()
        await processing_msg.edit_text(f"❌ Impossible de trouver {selected_model} dans la catégorie {category}.")
        return await start(query, context)

    except Exception as e:
        timeout_task.cancel()
        await processing_msg.edit_text(f"❌ Une erreur est survenue ({e}). Veuillez réessayer.")
        return CHOOSE_CATEGORY


async def respond_with_inventory_info(query_or_update, context, row, category, match_part):
    message = query_or_update.message if hasattr(query_or_update, 'message') else query_or_update.effective_message

    # Send "processing" message
    processing_message = await message.reply_text("⏳ Recherche en cours, veuillez patienter...")

    try:
        # Apply timeout
        async with asyncio.timeout(5):
            phone_model = row["Designation1"]
            available = row.get("QT", 0)
            price = row.get("PU", "-")

            matched_model_name = match_part.title()
            brand_match = next((b for b in KNOWN_BRANDS if b.lower() in phone_model.lower()), "")
            formatted_model = f"{brand_match.upper()} {matched_model_name}".strip()

            buttons = [
                [InlineKeyboardButton("📍 Voir l'emplacement du magasin", url=STORE_LOCATION_URL)],
                [InlineKeyboardButton("🔁 Démarrer une nouvelle recherche", callback_data="restart")]
            ]
            reply_markup = InlineKeyboardMarkup(buttons)

            # Delete "processing" once done
            await processing_message.delete()

            if isinstance(available, (int, float)) and available > 0:
                await log_request(category, formatted_model, True)
                await message.reply_text(
                    f"✅ {category.upper()} pour {formatted_model} est disponible.\n💵 Prix : {price} DA",
                    reply_markup=reply_markup
                )
            else:
                await log_request(category, formatted_model, False)
                await message.reply_text(
                    f"❌ Désolé, {category.lower()} pour {formatted_model} n'est pas disponible.",
                    reply_markup=reply_markup
                )
    except asyncio.TimeoutError:
        # Delete "processing" and send timeout error
        await processing_message.delete()
        await message.reply_text("⚠️ La recherche a pris trop de temps. Veuillez réessayer avec /start.")

    return CHOOSE_CATEGORY


async def restart_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Show processing message
    processing_msg = await query.message.reply_text("⏳ Processing...")

    async def timeout_fallback():
        await asyncio.sleep(5)
        if not context.user_data.get("response_sent"):
            buttons = [
                [InlineKeyboardButton("📱 LCD", callback_data="LCD")],
                [InlineKeyboardButton("🔋 Batterie", callback_data="Battery")],
                [InlineKeyboardButton("🔌 Connecteur", callback_data="Connector")],
                [InlineKeyboardButton("🧿 Glass", callback_data="Glass")],
                [InlineKeyboardButton("📦 Cover", callback_data="COVER")],
                [InlineKeyboardButton("🛠 Sersou", callback_data="SERSOU")]
            ]
            markup = InlineKeyboardMarkup(buttons)
            await query.message.reply_text(
                "❌ An error has occurred while processing your request. Please try again.",
                reply_markup=markup
            )

    timeout_task = asyncio.create_task(timeout_fallback())

    try:
        context.user_data.clear()
        inventory_cache["data"] = None

        timeout_task.cancel()
        await processing_msg.delete()
        context.user_data["response_sent"] = True

        await query.message.reply_text("🔁 Nouvelle recherche démarrée.")
        return await start(update, context)

    except Exception as e:
        timeout_task.cancel()
        await processing_msg.edit_text(f"❌ Une erreur est survenue ({e}). Veuillez réessayer.")
        return CHOOSE_CATEGORY

async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("🗕 Résumé d'aujourd'hui", callback_data="summary_today")],
        [InlineKeyboardButton("🗖 Résumé du mois", callback_data="summary_month")]
    ]
    markup = InlineKeyboardMarkup(buttons)
    await update.message.reply_text("Choisissez un type de résumé :", reply_markup=markup)

async def handle_summary_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data_type = query.data
    now = datetime.now()

    if data_type == "summary_today":
        date_str = now.strftime("%Y-%m-%d")
        log_file = os.path.join(LOG_DIR, f"{date_str}.log")
        summary = read_detailed_log_summary(log_file)
        await query.edit_message_text(f"🗕 Résumé du {date_str} :\n\n{summary}")

    elif data_type == "summary_month":
        month_str = now.strftime("%Y-%m")
        summaries = []
        for filename in os.listdir(LOG_DIR):
            if filename.startswith(month_str) and filename.endswith(".log"):
                log_file = os.path.join(LOG_DIR, filename)
                summaries.append(f"{filename[:-4]}:\n{read_detailed_log_summary(log_file)}\n")
        if summaries:
            await query.edit_message_text("🗖 Résumé du mois :\n\n" + "\n".join(summaries))
        else:
            await query.edit_message_text("📍 Aucun enregistrement pour ce mois.")

def read_detailed_log_summary(file_path):
    total = 0
    available = 0
    not_available = 0
    model_stats = Counter()
    not_available_models = []

    try:
        with open(file_path, 'r') as f:
            for line in f:
                total += 1
                parts = line.strip().split(" - ")
                if len(parts) < 3:
                    continue
                model = parts[1]
                status = parts[2]
                model_stats[model] += 1
                if "Available" in status:
                    available += 1
                else:
                    not_available += 1
                    not_available_models.append(model)

        top_models = model_stats.most_common(5)
        top_text = "\n".join([f"🔹 {model} ({count} demandes)" for model, count in top_models])
        unavailable_text = "\n".join([f"❌ {model}" for model in sorted(set(not_available_models))])

        return (
            f"🔎 Total: {total}\n"
            f"✅ Disponibles: {available}\n"
            f"❌ Non disponibles: {not_available}\n\n"
            f"📊 Top 5 modèles demandés :\n{top_text or 'Aucune donnée'}\n\n"
            f"🚫 Modèles non disponibles :\n{unavailable_text or 'Aucun'}"
        )
    except FileNotFoundError:
        return "Aucune donnée trouvée pour cette date."
